#https://www.datasciencearth.com/folium-ile-harita-gorsellestirme/

'''
Map() fonksiyonu içerisinde “tiles”
parametresi varsayılanı ”OpenStreetMap”
olarak görünüyor.
Bu parametre haritanın stilini değiştirmemizi sağlıyor.
Bunu Stamen Toner, CartoDB positron, Cartodb dark_matter,
Stamen Watercolor ya da Stamen Terrain olarak değiştirebiliyoruz.
width ve height parametreleriyle de boyutu ayarlıyoruz,
zoom_start ile de yaklaştırma işlemini yapıyoruz.
'''
import folium
from folium import GeoJson
import matplotlib as plt
import pandas as pd

#ek
import streamlit as st
from streamlit_folium import folium_static
st.set_page_config(layout ="wide")
import openpyxl as xl

#deneme
#Ek 3 I-O
import streamlit as st
import base64
import os

st.title("Bölge ve il bazında Girdi-Çıktı Tablosu Oluşturma")

options = st.multiselect(
    'Hangi illerin toplam girdi çıktı tablosunu oluşturmak istiyorsunuz?',
    ['Adana','Adıyaman','Afyon','Ağrı','Aksaray','Amasya','Ankara','Antalya',
     'Ardahan','Artvin','Aydın','Balıkesir','Bartın','Batman','Bayburt',
     'Bilecik','Bingöl','Bitlis','Bolu','Burdur','Bursa','Çanakkale','Çankırı',
     'Çorum','Denizli','Diyarbakır','Düzce','Edirne','Elazığ','Erzincan','Erzurum',
     'Eskişehir','Gaziantep','Giresun','Gümüşhane','Hakkari','Hatay','Iğdır','Isparta',
     'Istanbul','Izmir','Kahramanmaraş','Karabük','Karaman','Kars','Kastamonu','Kayseri',
     'Kırıkkale','Kırklareli','Kırşehir','Kilis','Kocaeli','Konya','Kütahya','Malatya','Manisa','Mardin',
     'Mersin','Muğla','Muş','Nevşehir','Niğde','Ordu','Osmaniye','Rize','Sakarya','Samsun',
     'Siirt','Sinop','Sivas','Şanlıurfa','Şırnak','Tekirdağ','Tokat','Trabzon','Tunceli',
     'Uşak','Van','Yalova','Yozgat','Zonguldak'],
    ['Istanbul', 'Ankara'])



#veriyi excelden kopyalama
#2. aşama
#Yüklenecek dosyaların oluşturulması.
#fonksiyon
def matris_toplam(il_a,il_b):
    a = il_a
    name_1  ="ILBAZINDA_2/il_bazinda_gsyh_ifk_a10_cari_deger_v5_" + a + ".xlsx"
    #Istanbul
    df = pd.read_excel(name_1,sheet_name="DB") 

    b = il_b
    name_2  ="ILBAZINDA_2/il_bazinda_gsyh_ifk_a10_cari_deger_v5_" + b + ".xlsx"

    df_a = pd.read_excel(name_2,sheet_name="DB")

    col_names  = df_a.columns

    df_total = df.loc[:,"A01":"T_Kullanim"] + df_a.loc[:,"A01":"T_Kullanim"]

    df_total = df.loc[:,"A01":"T_Kullanim"] + df_a.loc[:,"A01":"T_Kullanim"]

    df_total[['No', 'Bos', 'Kod_1', 'Bos_2', 'Aciklama',]]  =df[['No', 'Bos', 'Kod_1', 'Bos_2', 'Aciklama']]
    df_total= df_total[col_names]
    print(df_total)

    c = "Ara_Toplam"
    name_3  ="ILBAZINDA_2/il_bazinda_gsyh_ifk_a10_cari_deger_v5_" + c + ".xlsx"    
    df_total.to_excel(name_3,sheet_name="DB")
    return df_total

#Listede bulunan illerin kodlarıyla liste oluşturma
#sehirler = ["Ankara","Konya","Ordu","Kilis","Gaziantep","Istanbul"]
df_sehirler = pd.DataFrame(options, columns=['Ad'])

dim_1 = pd.read_excel("ILBAZINDA_2/dim_1.xlsx")

merge_1 = pd.merge(df_sehirler, dim_1,on="Ad",how="left")
liste_il = merge_1["Kod_1"].tolist()

print(merge_1)

#matris_toplam("TR100","TR310")

df_sonuc = pd.DataFrame()

squares = []
x = 0
#liste_il = ["TR100","TR310","TR510"]
for i in liste_il:
    x =x +1
    print("x deger")
    print(x)
    squares.append(i)
    print(squares)
    if len(squares) == 2:
        df_sum = matris_toplam(squares[0],squares[1])
        squares.clear()
        squares.append("Ara_Toplam")
        if x==len(liste_il):
            df_sum.to_excel("OUTPUT/sonuc_1.xlsx")
            
#3. aşama template oluşturma ve yüklenecek butona aktarma excel dosyayı

# opening the source excel file 
filename ="OUTPUT/sonuc_1.xlsx"
wb1 = xl.load_workbook(filename) 
ws1 = wb1.worksheets[0]

print(ws1)
  
# opening the destination excel file  
filename1 ="OUTPUT/template_1.xlsx"
wb2 = xl.load_workbook(filename1) 
ws2 = wb2.active 
  
# calculate total number of rows and  
# columns in source excel file 
mr = ws1.max_row 
mc = ws1.max_column 
  
# copying the cell values from source  
# excel file to destination excel file 
for i in range (2, 79): 
    for j in range (7, 84): 
        # reading cell value from source excel file 
        c = ws1.cell(row = i, column = j) 
  
        # writing the read value to destination excel file 
        ws2.cell(row = i+11, column = j).value = c.value 
  
# saving the destination excel file
filename2 ="OUTPUT/yayin_sonuc_1.xlsx"
wb2.save(str(filename2))

df = pd.read_excel("OUTPUT/yayin_sonuc_1.xlsx", sheet_name="YAYIN")


#Verileri Gönder
from io import BytesIO
from pyxlsb import open_workbook as open_xlsb

def to_excel(df):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    df.to_excel(writer, index=False, sheet_name='Sheet1')
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    format1 = workbook.add_format({'num_format': '0.00'}) 
    worksheet.set_column('A:A', None, format1)  
    #writer.save()
    writer.close()
    processed_data = output.getvalue()
    return processed_data

#Aşağıya Bak a3
#a3
df_xlsx = to_excel(df)

st.download_button(label='📰 Download Current Result',
                                data=df_xlsx ,
                                file_name= 'df_test.xlsx')


df_2 = pd.read_excel("OUTPUT/sonuc_1.xlsx")

st.write(df_2)




#Haritanın oluşturulması.
#4. aşama
''' Seçili olanlardan renklendirme alanlarını belirleme '''
df = pd.DataFrame(options)

data = pd.DataFrame(df[0].value_counts())
data= data.reset_index()
data.columns = ['Ad', 'Value']
print(data)


#devam

geo=r"tr-cities.json"
file = open(geo, encoding="utf8")
text = file.read()
 
m = folium.Map(location=[39,35],tiles="CartoDB positron",
               width="%100",weight="%100",zoom_start=6)
 
GeoJson(text).add_to(m)
#m.save("map_5.html")



folium.Choropleth(
    geo_data=text,
    data=data,
    columns=['Ad', "Value"],   
    legend_name='Seçili Bölgeler Toplam Yurtiçi Arz Kullanım Tablosu(2022)',
    key_on='feature.properties.name'
#‘feature.id’ ya da ‘feature.properties.statename’ de olabilir.
    ).add_to(m)

#sonuç
#m.save("map_2021.html")
#ek


folium_static(m, width=920, height=410)

print(options)
''' Seçili olanlardan renklendirme alanlarını belirleme '''

#Console'da görünsün diye var.
df = pd.DataFrame(options)

df_winners_company = pd.DataFrame(df[0].value_counts())
df_winners_company = df_winners_company.reset_index()
df_winners_company.columns = ['Ad', 'Value']
print(df_winners_company)



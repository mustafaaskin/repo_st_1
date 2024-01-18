#https://www.datasciencearth.com/folium-ile-harita-gorsellestirme/

'''
Map() fonksiyonu içerisinde “tiles”
parametresi varsayılanı ”OpenStreetMap”
olarak görünüyor.
Bu parametre haritanın stilini değiştirmemizi sağlıyor.
Bunu Stamen Toner, CartoDB positron, Cartodb dark_matter,
Stamen Watercolor ya da Stamen Terrain olarak değiştirebiliyoruz.
width ve height parametreleriyle de boyutu ayarlıyoruz,
zoom_start ile de yaklaştırma işlemini yapıyoruz.
'''
import folium
from folium import GeoJson
import matplotlib as plt
import pandas as pd

#ek
import streamlit as st
from streamlit_folium import folium_static
st.set_page_config(layout ="wide")
import openpyxl as xl

#deneme
#Ek 3 I-O
import streamlit as st
import base64
import os


tum_iller  = ['Adana','Adıyaman','Afyon','Ağrı','Aksaray','Amasya','Ankara','Antalya',
     'Ardahan','Artvin','Aydın','Balıkesir','Bartın','Batman','Bayburt',
     'Bilecik','Bingöl','Bitlis','Bolu','Burdur','Bursa','Çanakkale','Çankırı',
     'Çorum','Denizli','Diyarbakır','Düzce','Edirne','Elazığ','Erzincan','Erzurum',
     'Eskişehir','Gaziantep','Giresun','Gümüşhane','Hakkari','Hatay','Iğdır','Isparta',
     'Istanbul','Izmir','Kahramanmaraş','Karabük','Karaman','Kars','Kastamonu','Kayseri',
     'Kırıkkale','Kırklareli','Kırşehir','Kilis','Kocaeli','Konya','Kütahya','Malatya','Manisa','Mardin',
     'Mersin','Muğla','Muş','Nevşehir','Niğde','Ordu','Osmaniye','Rize','Sakarya','Samsun',
     'Siirt','Sinop','Sivas','Şanlıurfa','Şırnak','Tekirdağ','Tokat','Trabzon','Tunceli',
     'Uşak','Van','Yalova','Yozgat','Zonguldak']





#st.header(":blue[Create Input-Output Tables on Province and Regional Level] ")

st.markdown("""
<style>
.big-font {
    font-size:27px !important;
}
.upbig-font {
    font-size:29px !important;
   
}
.up-font {
    font-size:37px !important;
   
}
strong {
    color: #0068C9;
}
</style>
""", unsafe_allow_html=True)
st.markdown('<p class="up-font">İl Ayrıntısında, 2022 Türkiye Yurtiçi Girdi-Çıktı Tablosu / <strong> 2022 Turkiye Domestic Input-Output Table on Province Level</strong></p>',unsafe_allow_html=True)

st.markdown('<p class="upbig-font">İl ve Bölge Düzeyinde Yurtiçi Girdi-Çıktı Tablosu Oluşturun. / <strong> Create Input-Output Domestic Tables on Province and Regional Level</strong></p>',unsafe_allow_html=True)

#st.markdown('<p class="big-font">Hangi illerin toplam girdi çıktı tablosunu oluşturmak istiyorsunuz ?   / <strong>Which provinces do you want to choose and create total regional domestic input-output table ?</strong></p>', unsafe_allow_html=True)


options = st.multiselect(
    'İlleri Listeden Seçin / Choose Provinces From Dropdown List' ,
    ['Adana','Adıyaman','Afyon','Ağrı','Aksaray','Amasya','Ankara','Antalya',
     'Ardahan','Artvin','Aydın','Balıkesir','Bartın','Batman','Bayburt',
     'Bilecik','Bingöl','Bitlis','Bolu','Burdur','Bursa','Çanakkale','Çankırı',
     'Çorum','Denizli','Diyarbakır','Düzce','Edirne','Elazığ','Erzincan','Erzurum',
     'Eskişehir','Gaziantep','Giresun','Gümüşhane','Hakkari','Hatay','Iğdır','Isparta',
     'Istanbul','Izmir','Kahramanmaraş','Karabük','Karaman','Kars','Kastamonu','Kayseri',
     'Kırıkkale','Kırklareli','Kırşehir','Kilis','Kocaeli','Konya','Kütahya','Malatya','Manisa','Mardin',
     'Mersin','Muğla','Muş','Nevşehir','Niğde','Ordu','Osmaniye','Rize','Sakarya','Samsun',
     'Siirt','Sinop','Sivas','Şanlıurfa','Şırnak','Tekirdağ','Tokat','Trabzon','Tunceli',
     'Uşak','Van','Yalova','Yozgat','Zonguldak'],
    ['Istanbul', 'Ankara'])



#veriyi excelden kopyalama
#2. aşama
#Yüklenecek dosyaların oluşturulması.


#fonksiyon

   
name_a  ="ILBAZINDA_2/il_bazinda_gsyh_ifk_a10_cari_deger_v5_" + "TR" + ".xlsx"

df_a = pd.read_excel(name_a,sheet_name="DB")

def  tek_il_deger_al(il_a):
    a = il_a    
    name_1  ="ILBAZINDA_2/il_bazinda_gsyh_ifk_a10_cari_deger_v5_" + a + ".xlsx"
    #Istanbul
    df = pd.read_excel(name_1,sheet_name="DB")
    c = "Ara_Toplam"
    name_3  ="ILBAZINDA_2/il_bazinda_gsyh_ifk_a10_cari_deger_v5_" + c + ".xlsx"    
    df.to_excel(name_3,sheet_name="DB")
    return df

def matris_toplam(il_a,il_b):
    a = il_a
    name_1  ="ILBAZINDA_2/il_bazinda_gsyh_ifk_a10_cari_deger_v5_" + a + ".xlsx"
    #Istanbul
    df = pd.read_excel(name_1,sheet_name="DB") 

    b = il_b
    name_2  ="ILBAZINDA_2/il_bazinda_gsyh_ifk_a10_cari_deger_v5_" + b + ".xlsx"

    df_a = pd.read_excel(name_2,sheet_name="DB")

    col_names  = df_a.columns

    df_total = df.loc[:,"A01":"T_Kullanim"] + df_a.loc[:,"A01":"T_Kullanim"]

    df_total = df.loc[:,"A01":"T_Kullanim"] + df_a.loc[:,"A01":"T_Kullanim"]

    df_total[['No', 'Bos', 'Kod_1', 'Bos_2', 'Aciklama',]]  =df[['No', 'Bos', 'Kod_1', 'Bos_2', 'Aciklama']]
    df_total= df_total[col_names]
    print(df_total)

    c = "Ara_Toplam"
    name_3  ="ILBAZINDA_2/il_bazinda_gsyh_ifk_a10_cari_deger_v5_" + c + ".xlsx"    
    df_total.to_excel(name_3,sheet_name="DB")
    return df_total

#Listede bulunan illerin kodlarıyla liste oluşturma
#sehirler = ["Ankara","Konya","Ordu","Kilis","Gaziantep","Istanbul"]
df_sehirler = pd.DataFrame(options, columns=['Ad'])

dim_1 = pd.read_excel("ILBAZINDA_2/dim_1.xlsx")

merge_1 = pd.merge(df_sehirler, dim_1,on="Ad",how="left")
liste_il = merge_1["Kod_1"].tolist()

print(merge_1)

#matris_toplam("TR100","TR310")

df_sonuc = pd.DataFrame()

squares = []
x = 0
#liste_il = ["TR100","TR310","TR510"]
if len(options)>1:
    for i in liste_il:
        x =x +1
        print("x deger")
        print(x)
        squares.append(i)
        print(squares)
        if len(squares) == 2:
            df_sum = matris_toplam(squares[0],squares[1])
            squares.clear()
            squares.append("Ara_Toplam")
            if x==len(liste_il):
                df_sum.to_excel("OUTPUT/sonuc_1.xlsx")
elif len(options) == 1:
    for i in liste_il:
        squares.append(i)
        df_sum = tek_il_deger_al(i)
        squares.clear()
        squares.append("Ara_Toplam")
        df_sum.to_excel("OUTPUT/sonuc_1.xlsx")
else:
    df_a.to_excel("OUTPUT/sonuc_1.xlsx") 
    
        
            
#3. aşama template oluşturma ve yüklenecek butona aktarma excel dosyayı

# opening the source excel file 
filename ="OUTPUT/sonuc_1.xlsx"
wb1 = xl.load_workbook(filename) 
ws1 = wb1.worksheets[0]

print(ws1)
  
# opening the destination excel file  
filename1 ="OUTPUT/template_1.xlsx"
wb2 = xl.load_workbook(filename1) 
ws2 = wb2.active 
  
# calculate total number of rows and  
# columns in source excel file 
mr = ws1.max_row 
mc = ws1.max_column 
  
# copying the cell values from source  
# excel file to destination excel file 
for i in range (2, 79): 
    for j in range (7, 84): 
        # reading cell value from source excel file 
        c = ws1.cell(row = i, column = j) 
  
        # writing the read value to destination excel file 
        ws2.cell(row = i+11, column = j).value = c.value 
  
# saving the destination excel file
filename2 ="OUTPUT/yayin_sonuc_1.xlsx"
wb2.save(str(filename2))

df = pd.read_excel("OUTPUT/yayin_sonuc_1.xlsx", sheet_name="YAYIN",header=None)


#Verileri Gönder
from io import BytesIO
from pyxlsb import open_workbook as open_xlsb

def to_excel(df):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    df.to_excel(writer, index=False, sheet_name='Sheet1')
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    format1 = workbook.add_format({'num_format': '0.00'}) 
    worksheet.set_column('A:A', None, format1)  
    #writer.save()
    writer.close()
    processed_data = output.getvalue()
    return processed_data

#Aşağıya Bak a3
#a3
df_xlsx = to_excel(df)
print("processed data")
print(df_xlsx)

st.download_button(label='📰 Oluşan Tabloyu İndirin / Download Current Result',
                                data=df_xlsx ,
                                file_name= 'df_test.xlsx')





df_2 = pd.read_excel("OUTPUT/sonuc_1.xlsx", index_col = None)

df_2.rename(columns = {'Unnamed: 0':'A','Bos':'1', 'Bos_2':'2', 
                              'Bos_3':'3','Bos_4':'4'}, inplace = True) 
df_2 = df_2.drop(['A','1','2','3','4'], axis=1)
st.write(df_2)


print("Dikkat")
print(options)

#Haritanın oluşturulması.
#4. aşama
''' Harita. Seçili il ve/veya illerin alanları renklendirilmektedir.  --- Map. The area of Province and/or Provinces are colored.'''

if  len(options)>0:
    df = pd.DataFrame(options)
  
    data = pd.DataFrame(df[0].value_counts())
    data= data.reset_index()
    data.columns = ['Ad', 'Value']
    print(data)
    print("options")
    print(len(options))
    

else:
    data_2_a = {'Ad':['Istanbul','Ankara'],'Value':[0,0]}
    data_2 = pd.DataFrame(data_2_a)   

data_2_a = {'Ad':tum_iller,'Value':range(81)}
data_2 = pd.DataFrame(data_2_a) 
data_2['Value'] = 1
#devam

geo=r"tr-cities.json"
file = open(geo, encoding="utf8")
text = file.read()
 
m = folium.Map(location=[39,35],tiles="CartoDB positron",
               width="%100",weight="%100",zoom_start=6)
 
GeoJson(text).add_to(m)

n = folium.Map(location=[39,35],tiles="CartoDB positron",
               width="%100",weight="%100",zoom_start=6)
 
GeoJson(text).add_to(n)


#m.save("map_5.html")


if  len(options)>0:
    c = folium.Choropleth(
        geo_data=text,
        data=data,
        columns=['Ad', "Value"],   
        legend_name='Seçili Bölgeler Toplam Yurtiçi Arz Kullanım Tablosu(2022)',
        key_on='feature.properties.name'
    #‘feature.id’ ya da ‘feature.properties.statename’ de olabilir.
        ).add_to(m)
else:
    c = folium.Choropleth(
        geo_data=text,
        data=data_2,
        columns=['Ad', "Value"],   
        legend_name='Seçili Bölgeler Toplam Yurtiçi Arz Kullanım Tablosu(2022)',
        key_on='feature.properties.name'
    #‘feature.id’ ya da ‘feature.properties.statename’ de olabilir.
        ).add_to(m)


d = folium.Choropleth(
        geo_data=text,
        data=data_2,
        columns=['Ad', "Value"],   
        legend_name='Türkiye Toplam Yurtiçi Arz Kullanım Tablosu(2022)',
        key_on='feature.properties.name'
    #‘feature.id’ ya da ‘feature.properties.statename’ de olabilir.
        ).add_to(n)

#remove legend
#Harita 1
for key in c._children:
    if key.startswith('color_map'):
        del(c._children[key])

c.add_to(m)


#Harita 2
for key in d._children:
    if key.startswith('color_map'):
        del(d._children[key])

d.add_to(n)
#sonuç
#m.save("map_2021.html")
#ek

col1, col2 = st.columns(2)

with col1:
    
    if  len(options)>0:
        folium_static(m, width=920, height=410)
    else:
        #folium_static(m, width=920, height=410).empty()
        folium_static(n, width=920, height=410)  

print(options)


#Console'da görünsün diye var.
#df = pd.DataFrame(options)


#grafik ekle
#import numpy as np
#import itertools

df_gr_1 = df_2['Toplam_ARA_T']
print(df_gr_1)

df_gr_1  = df_gr_1.iloc[64:77]
df_gr_2 = df_2["Aciklama"].iloc[64:77]


print(len(df_gr_2))

df_g = pd.DataFrame(df_gr_1.values, index=df_gr_2)
print(df_gr_2)
#chart_data = pd.DataFrame(np.random.randn(20, 3), columns=["a", "b", "c"])
#colors  = list(itertools.repeat("#fd0",75))
with col2:    
    st.bar_chart(df_g)
    st.markdown("Katma Değer,İşletme Artığı, Çalışanlara Yapılan Ödemeler, Toplam Çıktı, Ara Tüketim, Toplam Kullanım Grafiği - (Toplam_ARA_T Sütunu)")
    st.markdown("Value Added, Operating Income, Compensation for Employees, Total Output, Intermediate Consumption, Total Use Graph - (Toplam_ARA_T Column)")



st.markdown("Dikkat: Sunulan veriler Resmi İstatistik Değildir. Bu uygulamada kullanılan veriler Mustafa AŞKIN tarafından ekonomik etki analizlerinde kullanılmak üzere \
oluşturulmuştur. Hesaplar oluşturulurken Türkiye İstatistik Kurumunun yayınladığı 2012 Yurtiçi Girdi-Çıktı Tablosu verilerinden,\
2022 Kurumsal Sektörel Hesaplar, Bölgesel Hesaplar, Yıllık ve Dönemsel Hesaplar verilerinden yararlanılmıştır.")
st.markdown("Attention: The data provided are not Official Statistics. The data provided in this application is estimated by MustafA AŞKIN \
for economic impact analysis. Data Source for estimates: 2012 Domestic Input-Output Table, \
2022 Institutional Sectoral Accounts, Regional Accounts, Annual  and Quarterly Accounts data which are disseminated by Turkish Statistical Institute.")


st.markdown('<a href="https://mustafaaskin.github.io/"> Bilgi ve İletişim İçin Linkler </a>',unsafe_allow_html=True)
st.markdown('<a href="https://mustafaaskin.github.io/"> Links for Information and Contact </a>',unsafe_allow_html=True)

